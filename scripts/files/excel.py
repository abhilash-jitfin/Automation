import os
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from .base import BaseFile


class ExcelFile(BaseFile):
    """Class representing an Excel file."""

    def split(self, output_dir: str, chunk_size: int) -> None:
        """
        Split the Excel file into chunks of the given size and save them to the output directory.

        :param output_dir: Directory where the chunks will be saved.
        :param chunk_size: Number of rows each chunk should contain.
        """
        if chunk_size <= 0:
            raise ValueError("chunk_size should be greater than 0")

        output_dir = Path(output_dir)

        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        total_rows = sheet.max_row
        headings = [cell.value for cell in sheet[1]]
        base_name_without_ext = os.path.splitext(os.path.basename(self.file_path))[0]
        print("\n")
        for i in range(2, total_rows, chunk_size):
            data = sheet[i : i + chunk_size] if i == 2 else sheet[i + 1 : i + chunk_size]
            df = pd.DataFrame(([cell.value for cell in row] for row in data), columns=headings)
            filepath = output_dir / f"{base_name_without_ext}_chunk_{i // chunk_size + 1}.xlsx"
            df.to_excel(filepath, index=False)
            print(filepath)
            # first_row_data = sheet[i+1]
            # first_row_values = [cell.value for cell in first_row_data]
            # print(f"First row content of file chunk_{i // chunk_size + 1}.xlsx: {first_row_values}\n")
            # last_row_data = sheet[i + chunk_size]
            # last_row_values = [cell.value for cell in last_row_data]
            # print(f"Last row content of file chunk_{i // chunk_size + 1}.xlsx: {last_row_values}")

    def read(self, sheet: Optional[str] = None, columns_to_read: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Read data from the Excel file.

        :param sheet: Name of the sheet to read.
        :param columns_to_read: List of columns to read.
        :return: DataFrame with the data from the file.
        """
        engine = self._get_engine_for_file_extension()
        df = pd.read_excel(
            self.file_path, sheet_name=sheet if sheet is not None else 0, usecols=columns_to_read, engine=engine
        )
        return df

    def _get_engine_for_file_extension(self) -> str:
        """
        Detect the file extension and return the appropriate engine for reading the Excel file.

        :return: Engine to be used for reading the Excel file.
        """
        _, extension = os.path.splitext(self.file_path)

        if extension.lower() == ".xlsx":
            return "openpyxl"
        elif extension.lower() == ".xls":
            return "xlrd"
        else:
            raise ValueError(f"Unsupported file extension: {extension}")

    def save(self, df: pd.DataFrame) -> None:
        """
        Save the given DataFrame to an Excel file.

        :param df: DataFrame to be saved.
        """
        df.to_excel(self.file_path, index=False)
